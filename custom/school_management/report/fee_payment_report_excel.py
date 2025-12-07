from odoo import models, fields, api
from odoo.exceptions import UserError

import io
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


class FeePaymentReportWizard(models.TransientModel):
    _name = 'fee.payment.report.wizard'
    _description = 'Fee Payment Report Wizard'

    date_from = fields.Date('Date From')
    date_to = fields.Date('Date To')
    payment_method = fields.Selection(
        [('cash', 'Cash'), ('bank_transfer', 'Bank Transfer'), ('cheque', 'Cheque'),
         ('card', 'Card'), ('other', 'Other')], string="Payment Method")
    state = fields.Selection(
        [('draft', 'Draft'), ('partially_paid', 'Partially Paid'),
         ('paid', 'Paid'), ('unpaid', 'Unpaid'), ('all', 'All')],
        default='all')

    def generate_fee_payment_excel_report(self):
        # Fetch records based on filters
        domain = []
        if self.date_from:
            domain.append(('payment_date', '>=', self.date_from))
        if self.date_to:
            domain.append(('payment_date', '<=', self.date_to))
        if self.state != 'all':
            domain.append(('state', '=', self.state))

        payments = self.env['fees.fees'].search(domain)

        if not payments:
            raise UserError("No records found for the selected criteria.")

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Fee Payment Report"

        # Define styles
        header_font = Font(bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        align_center = Alignment(horizontal='center', vertical='center')

        # Write headers
        headers = ['Student', 'Fee Type', 'Total Amount', 'Amount Paid',
                   'Amount Remaining', 'Status', 'Payment Method']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.alignment = align_center
            cell.border = border

        # Write data
        row = 2
        for payment in payments:
            ws.cell(row=row, column=1).value = payment.student_id.name
            ws.cell(row=row, column=2).value = payment.fee_type
            ws.cell(row=row, column=3).value = payment.total_amount
            ws.cell(row=row, column=4).value = payment.amount_paid
            ws.cell(row=row, column=5).value = payment.amount_remaining
            ws.cell(row=row, column=6).value = payment.state
            ws.cell(row=row, column=7).value = payment.payment_method
            row += 1

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        # Save the workbook to a binary stream
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Encode the file in base64
        excel_data = base64.b64encode(output.read())
        output.close()

        # Create Attachment
        attachment = self.env['ir.attachment'].create({
            'name': 'Fee_Payment_Report.xlsx',
            'type': 'binary',
            'datas': excel_data,
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Return action to download the file
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
